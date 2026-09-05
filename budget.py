# budget.py
import os
import sys
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import customtkinter as cck
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from budget_view import BudgetView
from charts_view import ChartsView
from config_manager import load_settings, save_settings
from crypto_utils import load_data, save_data
from themes import ThemeManager
from tools import FinancialToolsView

cck.set_appearance_mode("dark")
cck.set_default_color_theme("blue")

DEFAULT_CATEGORIES = ["Продукты", "Транспорт", "Кафе", "Жилье", "Здоровье", "Зарплата", "Другое"]


def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class BudgetApp(cck.CTk):
    def __init__(self):
        super().__init__()
        self.title("FinTrack Pro — Управление личными финансами")
        self.geometry("1080x700")
        self.minsize(960, 620)

        # Иконка приложения
        icon_path = get_resource_path("app_icon.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception:
                pass

        # 1. Загрузка конфигурации из settings.json
        cfg = load_settings()
        self.current_theme = cfg.get("theme", "dark")
        self.ui_scale = cfg.get("scale", 1.0)
        self.table_font_size = cfg.get("font_size", 11)
        self.currency_symbol = cfg.get("currency_symbol", "₽")
        self.currency_code = cfg.get("currency_code", "RUB")
        self.animations_enabled = cfg.get("animations_enabled", True)

        # 2. Загрузка базы транзакций
        raw = load_data()
        if isinstance(raw, dict):
            self.transactions = raw.get("transactions", [])
            self.categories = raw.get("categories", DEFAULT_CATEGORIES.copy())
        else:
            self.transactions = raw if isinstance(raw, list) else []
            self.categories = DEFAULT_CATEGORIES.copy()

        self.theme_manager = ThemeManager(self)

        # 3. Вкладки интерфейса
        self.tabs = cck.CTkTabview(
            self,
            corner_radius=12,
            fg_color="transparent",
            segmented_button_fg_color="#1E1E24",
            segmented_button_selected_color="#2563EB",
            segmented_button_selected_hover_color="#1D4ED8",
            segmented_button_unselected_color="#1E1E24",
            segmented_button_unselected_hover_color="#2A2A32",
            command=self._on_tab_change
        )
        self.tabs._segmented_button.configure(corner_radius=10)
        self.tabs.pack(fill="both", expand=True, padx=12, pady=(4, 10))

        tab_b = self.tabs.add("📊 Бюджет и операции")
        tab_c = self.tabs.add("📈 Аналитика и графики")
        tab_t = self.tabs.add("🛠 Калькуляторы и ЦБ РФ")

        # 4. Инициализация экранов
        self.budget_view = BudgetView(tab_b, self)
        self.charts_view = ChartsView(tab_c, lambda: self.transactions)
        FinancialToolsView(tab_t)

        # 5. Применение настроек
        self.theme_manager.apply(self.current_theme)
        self.set_table_font_size(self.table_font_size, save=False)
        self.budget_view.update_table(animate=False)

        if self.ui_scale != 1.0:
            self.after(50, lambda: cck.set_widget_scaling(self.ui_scale))

    def _on_tab_change(self):
        if self.tabs.get() == "📈 Аналитика и графики":
            self.charts_view.draw_charts()

    def export_to_csv(self):
        """Экспорт всей истории транзакций в профессиональный файл Excel (.xlsx)"""
        if not self.transactions:
            messagebox.showinfo("Экспорт", "История операций пуста. Нечего выгружать.")
            return

        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(desktop_path):
            desktop_path = os.path.abspath(".")

        # Открываем диалог сохранения файла с расширением .xlsx
        file_path = filedialog.asksaveasfilename(
            initialdir=desktop_path,
            defaultextension=".xlsx",
            filetypes=[("Книга Excel", "*.xlsx")],
            initialfile=f"fintrack_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            title="Выгрузить отчёт в Excel"
        )

        if not file_path:
            return

        try:
            # Создаем рабочую книгу openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "История транзакций"

            # Стилизация таблицы Excel
            font_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=10)
            fill_header = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # 1. Запись заголовков
            headers = ["Дата", "Тип операции", "Категория", f"Сумма ({self.currency_symbol})", "Описание"]
            ws.append(headers)

            # Применение стилей к шапке
            for col_num in range(1, 6):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_title
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            # 2. Запись данных транзакций
            for tx in self.transactions:
                row_data = [
                    tx.get("date", "-"),
                    tx.get("type", "-"),
                    tx.get("cat", "-"),
                    float(tx.get("amount", 0.0)),  # Сохраняем как число, чтобы Excel мог считать формулы
                    tx.get("desc", "-")
                ]
                ws.append(row_data)

                # Стилизация добавленной строки данных
                curr_row = ws.max_row
                ws.cell(row=curr_row, column=1).alignment = align_center  # Дата
                ws.cell(row=curr_row, column=2).alignment = align_center  # Тип
                ws.cell(row=curr_row, column=3).alignment = align_left  # Категория

                # Сумма (форматируем как числовой тип в ячейке)
                cell_amount = ws.cell(row=curr_row, column=4)
                cell_amount.alignment = align_right
                cell_amount.number_format = '#,##0.00'

                ws.cell(row=curr_row, column=5).alignment = align_left  # Описание

                # Добавляем границы и шрифт ко всем ячейкам строки
                for col_num in range(1, 6):
                    c = ws.cell(row=curr_row, column=col_num)
                    c.font = font_data
                    c.border = thin_border

            # 3. Автоматический расчёт ширины столбцов по длине содержимого
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Сохранение готового файла
            wb.save(file_path)
            messagebox.showinfo("Успех",
                                f"Отчёт успешно сохранён в формате Excel на Рабочий стол:\n{os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сгенерировать Excel файл:\n{e}")

    def set_table_font_size(self, size, save=True):
        self.table_font_size = size
        style = ttk.Style()
        row_height = int(size * 2.2)
        style.configure("Treeview", font=("Segoe UI", size), rowheight=row_height)
        style.configure("Treeview.Heading", font=("Segoe UI", size, "bold"))

        if hasattr(self, "budget_view") and hasattr(self.budget_view, "tree"):
            self.budget_view.tree.configure(style="Treeview")

        if save:
            self.save_app_settings()

    def set_ui_scale(self, scale):
        self.ui_scale = scale
        cck.set_widget_scaling(scale)
        self.save_app_settings()

    def set_currency(self, symbol, code):
        self.currency_symbol = symbol
        self.currency_code = code
        self.budget_view.update_table(animate=False)
        self.save_app_settings()

    def save_app_settings(self):
        save_settings({
            "theme": self.current_theme,
            "scale": self.ui_scale,
            "font_size": self.table_font_size,
            "currency_symbol": self.currency_symbol,
            "currency_code": self.currency_code,
            "animations_enabled": self.animations_enabled,
        })

    def save_all(self):
        save_data({
            "transactions": self.transactions,
            "categories": self.categories,
        })


if __name__ == "__main__":
    app = BudgetApp()
    app.mainloop()
